from openpyxl import load_workbook

###读取excel
#读取
#排序
def data():
    ordered_list_original =  [] #原始数据列表
    wb = load_workbook('学生成绩信息.xlsx')
    ws = wb.active
    for x in list(ws.iter_rows(max_col=2))[1:]:
        row = [data.value for data in x]
        ordered_list_original.append(row)
    # print(ordered_list_original)
    # 按照分数排序
    ordered_list_original =sorted(ordered_list_original,key=lambda x:x[1],reverse=True)
    # print(ordered_list_original)

    #制作名次字典
    student_ordered_dict = {}
    #注意并列排名问题的处理
    temp_compare = 0 #做一个比较用于查询是否有同分的问题
    rank = 0
    for x in ordered_list_original:
        # print(x[0],x[1])
        if x[1] == temp_compare :  #遇到同分数的， rank值不增长
            temp_compare = x[1] #当前值赋值给 temp用于下次比对
            student_ordered_dict[x[0]] = [x[1],rank] #向字典添加学员的信息
        else:
            temp_compare = x[1]
            rank = rank + 1 #遇到非同分的rank值增长
            student_ordered_dict[x[0]] = [x[1], rank]

    # print(student_ordered_dict)
    #--------以上完成基于学员信息的读取和名次排序 ------

    #--------下边我们把家长信息带进来-------

    parent_wb = load_workbook('家长信息.xlsx')
    parent_ws = parent_wb.active

    #调用家长的表 循环 ，然后与学生的list进行合并
    for x in list(parent_ws.iter_rows())[1:]:
        row = [data.value for data in x]
        #注意 row[0]学生名字 row[1]家长名字 row[2]称呼 row[3]手机 row[4]邮箱
        #两张表唯一的共有字段是学生的名字， 这个就是一个简单的vlookup

        temp_list = student_ordered_dict[row[0]]
        temp_list.extend([ row[1],row[2] ]) #注意这里要使用extend(可迭代对象) ，不能使用append，因为有好几个数据
        # print(temp_list)
        student_ordered_dict[ row[0] ] = temp_list

    # print(student_ordered_dict)

    return student_ordered_dict

# data()


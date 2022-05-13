import random
import qrcode
from openpyxl import load_workbook

from email.mime.multipart import MIMEMultipart  # 多个部件组装
from email.header import Header  # 邮件头部 标题 收件人等
from email.mime.text import MIMEText  # 邮件文字
from email.mime.application import MIMEApplication
from smtplib import SMTP_SSL  # 安全加密类似https/http


# qr制作
def qr_code(content):
    img = qrcode.make(content)
    filename = '%s.png' % content
    img.save('qr/%s' % filename)
    return filename


# 读取excel内容
def excel_read():
    user_info = []
    wb = load_workbook('students.xlsx')
    ws = wb.active
    for r in list(ws.iter_rows())[1:]:
        row = [data.value for data in r]  # 一行数一个list
        user_info.append(row)
    return user_info


# 邮件批量发送
def batch_mail(name,type,qr):
    # 先定义好发件人的邮箱地址和密码
    user = 'gzquse@163.com'
    pwd = 'KXEXSFCZWFEIASLE'
    # 目标收件人的邮箱地址
    target_mail = 'gzquse@163.com'
    # 然后定义好我们要使用的邮箱服务
    email_server = 'smtp.163.com'

    # todo 构造两个版本的 邮件内容
    # 构造 html内容VIP版本
    html_content_VIP = '''
    <html>
        <h1 style="color:white;margin-left:20px;background-color:blue"> 尊敬的VIP %s 你好！ </h1>
        <p style = "color:grey;font-size:20px">今天是您加入万门大学的第501天</p>
        <p>万门大学近期推出Python特训营项目，请点击下方链接关注我们</p>
        <a href = "https://www.wanmen.org">更多好课请关注万门大学</a>
        <p><img style = "width:100px" src='https://imgs.wanmen.org/3c8db5220d7cf0dc7b6ef81f524fea60.png'></p>
    </html> 
    ''' % (name)

    # 构造 html内容普通用户版本
    html_content_NOR = '''
    <html>
        <h1 style="color:lightblue;margin-left:20px"> 尊敬的 %s 你好 </h1>
        <p>万门大学近期推出Python特训营项目，请点击下方链接关注我们</p>
        <p>让知识改变人生！</p>
        <a href = "https://www.wanmen.org">更多好课请关注万门大学</a>
        <p><img style = "width:100px" src='https://imgs.wanmen.org/3c8db5220d7cf0dc7b6ef81f524fea60.png'></p>
        <h3>附件中有您的专属优惠二维码，现在加入万门大学VIP会员享受8折优惠</h3>
    </html>
    ''' % (name)

    # todo 这里要做一个用户的判断 ，确定使用哪个模板
    if type == '嘉宾':
        html_content = html_content_VIP
    elif type == '普通':
        html_content = html_content_NOR

    # todo 定义并读取附件
    attchment_file = MIMEApplication(open('qr/%s' % qr,'rb').read())
    attchment_file.add_header('Content-Disposition','attchment', filename='优惠码.png')

    # todo 建立邮件
    mail = MIMEMultipart()  # 实例化一个对象 -邮件部件组合对象
    # 组合邮件头部信息
    mail['Subject'] = Header('来自万门大学Python特训营的邀请函','utf-8')  # 邮件显示内容
    mail['From'] = Header('来自万门大学的邀请函！', 'utf-8')  # 邮件显示内容
    mail['To'] = target_mail  # 邮件显示内容

    # 组合邮件本体
    mail.attach(MIMEText(html_content,'html','utf-8'))  # 邮件显示内容
    # 组合邮件附件
    mail.attach(attchment_file)

    # todo 发送邮件

    # 邮件服务器调用发送
    smtp = SMTP_SSL(email_server)  # 实例化 对象 带入服务器地址
    smtp.login(user, pwd)  # 登录
    smtp.sendmail(from_addr=user, to_addrs=target_mail, msg=mail.as_string())  # 操作发送
    smtp.quit()  # 完成发送


if __name__ == '__main__':
    # 读取excle
    user_info = excel_read()
    # for循环每一条数据 ，并归类数据
    for info in user_info:  # [[0,],[2]]
        print(info)
        name = info[0]
        mobile = info[1]
        type = info[2]
        email = info[3]
        # 调用qr制作二维码 返回二维码文件地址
        qrfile = qr_code(mobile)
        # 调用批量邮件发送
        batch_mail(name=name, type=type, qr=qrfile)
        print('发送成功！')

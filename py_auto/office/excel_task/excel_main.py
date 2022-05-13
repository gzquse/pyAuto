from openpyxl import load_workbook, Workbook
import os
from openpyxl.styles import Font, colors, Alignment, Border, Side, Color, Fill

# excel读取出来的全部数据
main_data = {}

# 分公司各自的总成本
individual_cost = {}

# 公司总工资成本
total = 0

# 平均工资
avg_cost_per = {}

# 平均产出/总产出
individual_output = {}
avg_output = {}

# 岗位平均工资
avg_pos_cost = {}

# 盈利能力
profit_company = {}

for file in os.listdir('maindata'):
    # print(file)
    wb = load_workbook(os.path.join('maindata', file))
    ws = wb.worksheets[0]
    temp = []
    for r in list(ws.iter_rows())[1:]:
        row = [data.value for data in r]  # 一行数一个list
        temp.append(row)
        main_data[file.split('.')[0]] = temp

# print(main_data)

# 计算分公司工资成本
for x, y in main_data.items():
    temp = []
    for cost in y:
        temp.append(cost[4])
    individual_cost[x] = sum(temp)
# print(individual_cost)

# 总工资成本
for _, cost in individual_cost.items():
    total = total + cost
# print(total)
# #

# ------------计算 各个公司的平均工资 ---------

for company, staff in main_data.items():
    for company_, cost in individual_cost.items():
        if company == company_:
            print(company, cost, len(staff))
            avg_cost_per[company] = cost / len(staff)
# print(avg_cost_per)


# ------------计算 各个公司的平均产出 ---------
# --先找出来每个公司的总产出
for x, y in main_data.items():
    temp = []
    for cost in y:
        temp.append(cost[5])
    individual_output[x] = sum(temp)
# print(individual_output)


for company, staff in main_data.items():
    for company_, cost in individual_output.items():
        if company == company_:
            avg_output[company] = cost / len(staff)
# print(avg_output)

# ----统计出来各个岗位的平均工资-----
# ---这里注意---- 大维度发生了变化 以岗位为维度
position_list = []
position_dict = {}
# 准备基础数据
for company, info in main_data.items():
    for x in info:
        position_list.append(x)
# print(position_list)
# 重新排列数据
print("**" * 80)
for pos in position_list:
    if position_dict.get(pos[1]) is not None:
        current = position_dict[pos[1]]
        current.append(pos)
        position_dict[pos[1]] = current
    else:
        position_dict[pos[1]] = [pos]

# print(position_dict) #岗位维度数据获取
# 计算岗位平均工资

for pos, staff in position_dict.items():
    temp = 0
    for i in staff:
        # print(i[4])
        temp = temp + i[4]
        avg_pos_cost[pos] = round(temp / len(pos))
    # print(temp)
# print(avg_pos_cost)


# ---统计出来平均盈利能力最好的公司-----

for company, output in avg_output.items():
    for company_, cost in avg_cost_per.items():
        if company == company_:
            profit = output - cost
            profit_company[company] = profit
# 排顺序
BEST = sorted(profit_company.items(), key=lambda d: d[1], reverse=True)
print(profit_company)
# print(profit_company)
print(dict(BEST))  # 盈利

# *****写入Excel******


wb = Workbook()
ws = wb.create_sheet()
ws.title = 'Myreport'
ws.column_dimensions['F'].width = 18.0  # 列宽设置
ws.column_dimensions['B'].width = 18.0  # 列宽设置

# 设置
# "A1:B3"
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
ws.cell(1, 1).value = '工资报告'
ws.cell(1, 1).font = Font(name='黑体', size=30, )
ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')

ws.merge_cells("B4:C4")
ws.cell(4, 2).value = '公司总工资成本'

border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))
# 注意这里有坑
ws.cell(4, 3).border = border
ws.cell(4, 2).border = border

# 对齐
ws.merge_cells('B5:C5')
ws.cell(5, 2).alignment = Alignment(horizontal='center', vertical='center')
ws['B5'].value = total

# 分公司公司的工资
row = 4
for company, cost in individual_cost.items():
    ws.cell(row, 6).value = company
    ws.cell(row, 7).value = cost
    row += 1

# 岗位平均工资
row = 10
for company, cost in avg_pos_cost.items():
    ws.cell(row, 2).value = company
    ws.cell(row, 3).value = cost
    row += 1

# 平均工资成本
row = 10
for company, cost in avg_cost_per.items():
    ws.cell(row, 6).value = company
    ws.cell(row, 7).value = cost
    row += 1

# 平均产出
row = 16
for company, cost in avg_output.items():
    ws.cell(row, 6).value = company
    ws.cell(row, 7).value = cost
    row += 1

# 平均盈利能力最好的公司

ws['B25'].value = BEST[0][0]
ws['C25'].value = BEST[0][1]
ws['C25'].font = Font(size=20, underline='single')

wb.save('abc.xlsx')
